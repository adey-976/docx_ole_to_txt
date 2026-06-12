"""
Extract full text from a .docx file including embedded files (.xlsx, .xlsm, .csv,
.msg, .eml, .txt, .docx, .pdf) at their exact positions within the document.

Usage:
    python3 extract_docx.py <path_to_docx> > <path_to_save_output_txt>
"""

import zipfile
import os
import struct
import email
from xml.etree import ElementTree as ET
from io import BytesIO

import docx
import openpyxl
import olefile


# XML namespaces used in .docx files
NSMAP = {
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'o': 'urn:schemas-microsoft-com:office:office',
    'v': 'urn:schemas-microsoft-com:vml',
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}

for prefix, uri in NSMAP.items():
    ET.register_namespace(prefix, uri)

# Maps OLE ProgID to file extension for identifying embedded file types
PROGID_EXT_MAP = {
    'Word.Document.12': '.docx',
    'Word.Document.8': '.doc',
    'Excel.Sheet.12': '.xlsx',
    'Excel.Sheet.8': '.xls',
    'Excel.SheetMacroEnabled.12': '.xlsm',
    'Acrobat.Document.DC': '.pdf',
    'Acrobat.Document': '.pdf',
    'Package': '',
}


def extract_xlsx_text(data: bytes) -> str:
    """Read all sheets and rows from an Excel file."""
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        lines.append(f"  [Sheet: {sheet}]")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = '\t'.join(str(c) if c is not None else '' for c in row)
            if row_text.strip():
                lines.append(f"    {row_text}")
    wb.close()
    return '\n'.join(lines)


def extract_csv_text(data: bytes) -> str:
    """Decode CSV bytes to string."""
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin-1')


def extract_txt_text(data: bytes) -> str:
    """Decode plain text bytes to string."""
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin-1')


def extract_eml_text(data: bytes) -> str:
    """Parse .eml file and extract headers + body."""
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        text = data.decode('latin-1')
    try:
        msg = email.message_from_string(text)
        parts = []
        if msg['Subject']:
            parts.append(f"  Subject: {msg['Subject']}")
        if msg['From']:
            parts.append(f"  From: {msg['From']}")
        if msg['To']:
            parts.append(f"  To: {msg['To']}")
        if msg['Date']:
            parts.append(f"  Date: {msg['Date']}")
        # Get the email body content
        body = msg.get_payload(decode=True)
        if body:
            parts.append(f"  Body: {body.decode('utf-8', errors='replace')}")
        elif isinstance(msg.get_payload(), str):
            parts.append(f"  Body: {msg.get_payload()}")
        return '\n'.join(parts)
    except Exception:
        return text


def extract_msg_text(data: bytes) -> str:
    """Parse Outlook .msg file using extract_msg library."""
    try:
        import extract_msg
        msg = extract_msg.openMsg(BytesIO(data))
        parts = []
        if msg.subject:
            parts.append(f"  Subject: {msg.subject}")
        if msg.sender:
            parts.append(f"  From: {msg.sender}")
        if msg.to:
            parts.append(f"  To: {msg.to}")
        if msg.body:
            parts.append(f"  Body: {msg.body}")
        return '\n'.join(parts)
    except Exception:
        # Fallback: treat as plain text
        try:
            text = data.decode('utf-8')
        except UnicodeDecodeError:
            text = data.decode('latin-1')
        return text


def extract_pdf_text(data: bytes) -> str:
    """Extract text from PDF using PyPDF2 or pdfplumber."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(BytesIO(data))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                pages.append(f"  [Page {i+1}]")
                pages.append(f"    {text.strip()}")
        return '\n'.join(pages) if pages else "  [Empty PDF]"
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(data)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"  [Page {i+1}]")
                        pages.append(f"    {text.strip()}")
                return '\n'.join(pages) if pages else "  [Empty PDF]"
        except ImportError:
            return "  [PDF detected but PyPDF2/pdfplumber not installed. Install with: pip install PyPDF2]"


def extract_embedded_docx_text(data: bytes) -> str:
    """Extract paragraph text from an embedded .docx file."""
    doc = docx.Document(BytesIO(data))
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())


def extract_ole_embedded(data: bytes, prog_id: str = ''):
    """
    Extract the embedded file from an OLE compound document.
    Returns (filename, file_bytes) or (None, None).
    """
    try:
        ole = olefile.OleFileIO(BytesIO(data))

        # CONTENTS stream: used by Adobe Acrobat for PDF embeddings
        if ole.exists('CONTENTS'):
            stream = ole.openstream('CONTENTS')
            file_data = stream.read()
            ole.close()
            ext = PROGID_EXT_MAP.get(prog_id, '')
            filename = f'embedded{ext}' if ext else 'embedded_contents'
            return filename, file_data

        # Ole10Native stream: used by Package-style embeddings (.eml, .txt, etc.)
        if ole.exists('\x01Ole10Native'):
            stream = ole.openstream('\x01Ole10Native')
            content = stream.read()
            ole.close()
            return parse_ole10native(content)

        # Package stream: another way files can be stored
        if ole.exists('Package'):
            stream = ole.openstream('Package')
            file_data = stream.read()
            ole.close()
            return 'unknown_package', file_data

        ole.close()
    except Exception:
        pass
    return None, None


def parse_ole10native(content: bytes):
    """
    Parse Ole10Native stream to extract the original filename and file data.
    Handles both Package format and Simple format.
    """
    if len(content) < 8:
        return None, None

    idx = 0

    # First 4 bytes: total size of the stream data
    total_size = struct.unpack('<I', content[idx:idx+4])[0]
    idx += 4

    # Next 2 bytes: flags
    flags = struct.unpack('<H', content[idx:idx+2])[0]
    idx += 2

    # Null-terminated label string (usually the filename)
    end = content.index(b'\x00', idx)
    label = content[idx:end].decode('latin-1')
    idx = end + 1

    # Null-terminated source path string
    end = content.index(b'\x00', idx)
    src_path = content[idx:end].decode('latin-1')
    idx = end + 1

    # Use label as filename, fall back to basename of source path
    filename = label if label else os.path.basename(src_path)

    if idx + 4 <= len(content):
        next_dword = struct.unpack('<I', content[idx:idx+4])[0]

        if next_dword == 0x00030000:
            # Package format: reserved DWORD, then temp path length, temp path, data size, data
            idx += 4
            temp_len = struct.unpack('<I', content[idx:idx+4])[0]
            idx += 4
            idx += temp_len  # skip temp path bytes
            if idx + 4 <= len(content):
                data_size = struct.unpack('<I', content[idx:idx+4])[0]
                idx += 4
                file_data = content[idx:idx+data_size]
                return filename, file_data
        else:
            # Simple format: two more null-terminated path strings, then data size + data
            try:
                end = content.index(b'\x00', idx)
                idx = end + 1
                end = content.index(b'\x00', idx)
                idx = end + 1
                if idx + 4 <= len(content):
                    data_size = struct.unpack('<I', content[idx:idx+4])[0]
                    idx += 4
                    file_data = content[idx:idx+data_size]
                    return filename, file_data
            except ValueError:
                pass

    return None, None


def guess_extension_from_data(data: bytes) -> str:
    """Guess file extension by checking magic bytes at the start of the file."""
    if data[:4] == b'PK\x03\x04':
        # ZIP-based format — check internal structure to identify type
        try:
            zf = zipfile.ZipFile(BytesIO(data))
            names = zf.namelist()
            if any('word/' in n for n in names):
                return '.docx'
            elif any('xl/' in n for n in names):
                return '.xlsx'
            elif any('ppt/' in n for n in names):
                return '.pptx'
            zf.close()
        except Exception:
            pass
        return '.zip'
    elif data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return '.ole'
    elif data[:5] == b'%PDF-':
        return '.pdf'
    elif data[:5] == b'MIME-' or data[:9] == b'Received:' or data[:5] == b'From:':
        return '.eml'
    return ''


def read_embedded_file(filename: str, data: bytes) -> str:
    """Route embedded file to the correct text extractor based on extension."""
    ext = os.path.splitext(filename)[1].lower()
    if not ext:
        ext = guess_extension_from_data(data)
    if ext in ('.xlsx', '.xlsm'):
        return extract_xlsx_text(data)
    elif ext == '.csv':
        return extract_csv_text(data)
    elif ext == '.txt':
        return extract_txt_text(data)
    elif ext == '.eml':
        return extract_eml_text(data)
    elif ext == '.msg':
        return extract_msg_text(data)
    elif ext == '.docx':
        return extract_embedded_docx_text(data)
    elif ext == '.pdf':
        return extract_pdf_text(data)
    else:
        return f"  [Unsupported format: {ext or 'unknown'}]"


def extract_document(docx_path: str) -> str:
    """Main function: extract all text from docx including embedded files."""
    zf = zipfile.ZipFile(docx_path)

    # Parse the relationships file to map relationship IDs to file targets and types
    rels_xml = zf.read('word/_rels/document.xml.rels')
    rels_tree = ET.fromstring(rels_xml)
    rels = {}
    rel_types = {}
    for rel in rels_tree:
        rid = rel.get('{http://schemas.openxmlformats.org/package/2006/relationships}Id')
        if rid is None:
            rid = rel.get('Id')
        target = rel.get('Target')
        rel_type = rel.get('Type', '')
        if rid and target:
            rels[rid] = target
            rel_types[rid] = rel_type

    # Parse the main document XML
    doc_xml = zf.read('word/document.xml')
    tree = ET.fromstring(doc_xml)
    body = tree.find(f'.//{{{NSMAP["w"]}}}body')

    output = []

    def process_paragraph(element):
        """Process a single paragraph element, extracting text and embedded objects."""
        results = []
        para_text = ''.join(
            node.text or ''
            for node in element.iter(f'{{{NSMAP["w"]}}}t')
        )

        embedded_found = False

        for obj in element.iter(f'{{{NSMAP["o"]}}}OLEObject'):
            rid = obj.get(f'{{{NSMAP["r"]}}}id')
            if not rid or rid not in rels:
                continue

            target = rels[rid]
            rel_type = rel_types.get(rid, '')
            prog_id = obj.get('ProgID', '')
            embed_path = f'word/{target}' if not target.startswith('/') else target.lstrip('/')

            try:
                embed_data = zf.read(embed_path)
            except KeyError:
                continue

            filename = os.path.basename(target)
            file_data = None

            if 'package' in rel_type.lower():
                file_data = embed_data
                ext = os.path.splitext(filename)[1].lower()
                if not ext or ext == '.bin':
                    prog_ext = PROGID_EXT_MAP.get(prog_id, '')
                    if prog_ext:
                        filename = os.path.splitext(filename)[0] + prog_ext

            elif 'oleobject' in rel_type.lower().replace('/', ''):
                ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
                if ole_filename and ole_data:
                    filename = ole_filename
                    file_data = ole_data
                else:
                    file_data = embed_data

            else:
                ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
                if ole_filename and ole_data:
                    filename = ole_filename
                    file_data = ole_data
                else:
                    file_data = embed_data

            if file_data:
                embedded_found = True
                if para_text.strip():
                    results.append(para_text)
                    para_text = ''
                results.append(f"\n--- Embedded File: {filename} ---")
                results.append(read_embedded_file(filename, file_data))
                results.append(f"--- End: {filename} ---\n")

        if not embedded_found and para_text.strip():
            results.append(para_text)

        return results

    # Walk through each top-level element in the document body
    for element in body:
        tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag

        if tag == 'p':
            output.extend(process_paragraph(element))

        elif tag == 'tbl':
            for row in element.iter(f'{{{NSMAP["w"]}}}tr'):
                cells = []
                row_has_embed = False
                for cell in row.iter(f'{{{NSMAP["w"]}}}tc'):
                    # Check if this cell contains any OLE objects
                    ole_in_cell = cell.findall(f'.//{{{NSMAP["o"]}}}OLEObject')
                    if ole_in_cell:
                        # Flush any accumulated cells as a row first
                        if cells:
                            output.append(' | '.join(cells))
                            cells = []
                        row_has_embed = True
                        # Process each paragraph in this cell individually
                        for para in cell.iter(f'{{{NSMAP["w"]}}}p'):
                            output.extend(process_paragraph(para))
                    else:
                        cell_text = ''.join(
                            node.text or ''
                            for node in cell.iter(f'{{{NSMAP["w"]}}}t')
                        )
                        cells.append(cell_text)
                if cells:
                    output.append(' | '.join(cells))

    zf.close()
    return '\n'.join(output)


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 extract_docx.py <path_to_docx> > <path_to_save_output_txt>")
        sys.exit(1)
    result = extract_document(sys.argv[1])
    print(result)
